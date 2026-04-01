#!/usr/bin/env python3
"""
parse_polyspace.py  -  Phase 6a: Parse QAC / Polyspace Excel warning report.

Supports two formats (auto-detected from headers):

  QAC format (real client data - Excel 2 Sheet 1):
    SL.No | File | Line No | Severity | Rule/Directives No. | Warning No
    Description | Line in code | judgment

  Mock/Polyspace format (legacy testing):
    Tool Name | Warning ID | Category | Checker Name | Rule ID | Severity
    Message   | File Path  | Line Start | Line End | Function Name

For each warning:
  - warning_to_rule DB is checked  -> gets MISRA Rule + M/R/A category
  - patch_library DB is checked    -> gets approved After Code fix (Fix 1)
  - LLM generates Fix 2+
  - Source context from .c file OR pasted code
"""

from __future__ import annotations

import json
import sqlite3
import sys
from pathlib import Path
from typing import Any, Dict, List, Optional

import openpyxl

CONTEXT_LINES = 10
SEVERITY_RANK = {
    "High": 3, "Medium": 2, "Low": 1,
    "Mandatory": 3, "Required": 2, "Advisory": 1,
}
SEVERITY_NORM = {
    "mandatory": "Mandatory", "required": "Required", "advisory": "Advisory",
    "high": "High", "medium": "Medium", "low": "Low",
    "m": "Mandatory", "r": "Required", "a": "Advisory",
}
QAC_COLS  = {"file", "line no", "warning no", "description"}
MOCK_COLS = {"warning id", "rule id", "message", "file path"}
NEW_MOCK_COLS = {"warning number", "rule id", "message", "file name"}


def _db_path() -> Path:
    return Path(__file__).resolve().parent.parent.parent / "data" / "knowledge" / "patch_library.db"


def _db_lookup(table: str, warning_no: str) -> Optional[Dict]:
    p = _db_path()
    if not p.exists():
        return None
    try:
        conn = sqlite3.connect(str(p))
        conn.row_factory = sqlite3.Row
        row = conn.execute(
            f"SELECT * FROM {table} WHERE warning_no=? LIMIT 1",
            (str(warning_no).strip(),)
        ).fetchone()
        conn.close()
        return dict(row) if row else None
    except Exception:
        return None


def _norm_sev(raw: str) -> str:
    return SEVERITY_NORM.get(str(raw).strip().lower(), str(raw).strip() or "Advisory")


def extract_source_context(
    source_dir: Path,
    file_path: str,
    line_no: Optional[int],
    inline_code: str = "",
    context: int = CONTEXT_LINES,
) -> Dict[str, Any]:
    """
    Build source context for the flagged code viewer.
    Priority: uploaded .c file > inline_code from Excel > pasted_code.c
    """
    src_file = None
    candidates = []
    if file_path:
        candidates = [source_dir / file_path, source_dir / Path(file_path).name]
    src_file = next((p for p in candidates if p.exists()), None)

    if src_file is None and file_path:
        c_files = list(source_dir.glob("*.c")) + list(source_dir.glob("*.h"))
        fp_stem = Path(file_path).stem.lower()
        src_file = next((f for f in c_files if f.stem.lower() == fp_stem), None)

    # Inline code from Excel "Line in code" column
    if src_file is None and inline_code:
        lines_from_excel = inline_code.splitlines()
        # Wrap inline snippet with line numbers starting at line_no or 1
        start = line_no or 1
        numbered = []
        for i, ln in enumerate(lines_from_excel):
            lineno = start + i
            marker = ">>>" if i == 0 else "   "
            numbered.append(f"{marker} {lineno:4d}  {ln}")
        return {
            "file": file_path or "inline",
            "found": True,
            "context_text": "\n".join(numbered),
            "context_start_line": start,
            "context_end_line": start + len(lines_from_excel) - 1,
            "flagged_lines": [start],
            "source": "inline",
        }

    if src_file is None:
        pasted = source_dir / "pasted_code.c"
        if pasted.exists():
            src_file = pasted

    if src_file is None:
        return {
            "file": file_path or "unknown", "found": False,
            "context_text": f"[Source file not found: {file_path}]",
            "context_start_line": None, "context_end_line": None,
            "flagged_lines": [],
        }

    all_lines = src_file.read_text(encoding="utf-8", errors="replace").splitlines()
    total = len(all_lines)

    if not line_no:
        ctx_start, ctx_end = 1, min(total, 40)
        ls = le = 0
        flagged = []
    else:
        ls = le = int(line_no)
        if ls <= total and not all_lines[ls - 1].strip():
            for look in range(ls + 1, min(le + 4, total + 1)):
                if all_lines[look - 1].strip():
                    ls = le = look
                    break
        ctx_start = max(1, ls - context)
        ctx_end   = min(total, le + context)
        flagged   = list(range(ls, le + 1))

    numbered = []
    for i in range(ctx_start - 1, ctx_end):
        lineno = i + 1
        marker = ">>>" if flagged and ls <= lineno <= le else "   "
        numbered.append(f"{marker} {lineno:4d}  {all_lines[i]}")

    return {
        "file": src_file.name, "found": True,
        "context_text": "\n".join(numbered),
        "context_start_line": ctx_start, "context_end_line": ctx_end,
        "flagged_lines": flagged,
    }


def _detect_format(headers: List[str]) -> str:
    lower = {h.lower().strip() for h in headers if h}
    if len(QAC_COLS & lower) >= 3:
        return "qac"
    # Accept either old mock format (warning id / file path) or new format (warning number / file name)
    mock_score     = len(MOCK_COLS & lower)
    new_mock_score = len(NEW_MOCK_COLS & lower)
    if max(mock_score, new_mock_score) >= 3:
        return "mock"
    return "qac" if len(QAC_COLS & lower) >= len(MOCK_COLS & lower) else "mock"


def _build_idx(headers: List[str]) -> Dict[str, int]:
    return {h.strip().lower(): i for i, h in enumerate(headers) if h}


def _get(row, idx: Dict[str, int], *keys) -> str:
    for k in keys:
        i = idx.get(k.lower())
        if i is not None and i < len(row) and row[i] is not None:
            return str(row[i]).strip()
    return ""


def _parse_qac(rows, headers, source_dir, uploaded_stems) -> List[Dict]:
    idx = _build_idx(headers)
    warnings = []
    skipped = 0

    for row_num, row in enumerate(rows[1:], start=2):
        if not any(row):
            continue

        file_path   = _get(row, idx, "file")
        line_no_raw = _get(row, idx, "line no", "line no.")
        warning_no  = _get(row, idx, "warning no", "warning no.")
        description = _get(row, idx, "description")
        severity_raw= _get(row, idx, "severity")
        rule_raw    = _get(row, idx, "rule / directives no.", "rule/directives no.", "rule no")
        inline_code = _get(row, idx, "line in code", "line_in_code", "code")

        if not file_path and not warning_no:
            continue

        excel_stem = Path(file_path).stem.lower() if file_path else ""
        if uploaded_stems and excel_stem and excel_stem not in uploaded_stems:
            skipped += 1
            continue

        severity = _norm_sev(severity_raw)
        line_no  = int(line_no_raw) if line_no_raw and str(line_no_raw).strip().isdigit() else None

        rule_map  = _db_lookup("warning_to_rule", warning_no) if warning_no else None
        patch_rec = _db_lookup("patch_library",   warning_no) if warning_no else None

        rule_id        = rule_map.get("misra_rule", rule_raw) if rule_map else rule_raw
        misra_category = rule_map.get("misra_category", severity) if rule_map else severity

        approved_fix = None
        if patch_rec and patch_rec.get("after_code"):
            approved_fix = patch_rec["after_code"]
        elif patch_rec and patch_rec.get("code_patch"):
            approved_fix = patch_rec["code_patch"]

        sl_no = _get(row, idx, "sl.no", "sl no", "slno")
        wid   = f"W{int(sl_no):04d}" if sl_no and str(sl_no).strip().isdigit() else f"W{row_num:04d}"

        source_ctx = extract_source_context(source_dir, file_path, line_no, inline_code)

        warnings.append({
            "warning_id":      wid,
            "warning_no":      warning_no,
            "rule_id":         rule_id,
            "rule_raw":        rule_raw,
            "severity":        severity,
            "severity_rank":   SEVERITY_RANK.get(severity, 1),
            "misra_category":  misra_category,
            "message":         description,
            "file_path":       file_path,
            "line_start":      line_no,
            "line_end":        line_no,
            "inline_code":     inline_code,
            "approved_fix":    approved_fix,
            "has_approved_fix": bool(approved_fix),
            "client_accepted": not bool(approved_fix),
            "source_context":  source_ctx,
        })

    if skipped:
        print(f"  Skipped {skipped} warnings - source files not uploaded")

    warnings.sort(key=lambda w: (-w["severity_rank"], w["file_path"], w["line_start"] or 0))
    return warnings


def _parse_mock(rows, headers, source_dir, uploaded_stems) -> List[Dict]:
    idx = _build_idx(headers)
    warnings = []
    skipped = 0

    for row in rows[1:]:
        if not any(row):
            continue

        file_path  = _get(row, idx, "file path", "file name", "filename", "file")
        ls_raw     = _get(row, idx, "line start", "line no", "line")
        le_raw     = _get(row, idx, "line end",   "line no", "line")
        warning_no = _get(row, idx, "warning id", "warning number", "warning no", "warning_no")
        severity   = _get(row, idx, "severity") or "Medium"
        rule_raw   = _get(row, idx, "rule id",   "rule/directives no.", "rule no", "rule")

        excel_stem = Path(file_path).stem.lower() if file_path else ""
        if uploaded_stems and excel_stem and excel_stem not in uploaded_stems:
            skipped += 1
            continue

        rule_map  = _db_lookup("warning_to_rule", warning_no) if warning_no else None
        patch_rec = _db_lookup("patch_library",   warning_no) if warning_no else None

        rule_id        = rule_map.get("misra_rule", rule_raw) if rule_map else rule_raw
        misra_category = rule_map.get("misra_category", severity) if rule_map else severity
        approved_fix   = patch_rec.get("after_code") if patch_rec else None

        line_no = int(ls_raw) if ls_raw and str(ls_raw).strip().isdigit() else None
        source_ctx = extract_source_context(source_dir, file_path, line_no)

        warnings.append({
            "warning_id":      warning_no,
            "warning_no":      warning_no,
            "tool_name":       _get(row, idx, "tool name"),
            "category":        _get(row, idx, "category"),
            "checker_name":    _get(row, idx, "checker name"),
            "rule_id":         rule_id,
            "severity":        severity,
            "severity_rank":   SEVERITY_RANK.get(severity, 1),
            "misra_category":  misra_category,
            "message":         _get(row, idx, "message"),
            "file_path":       file_path,
            "line_start":      line_no,
            "line_end":        int(le_raw) if le_raw and str(le_raw).strip().isdigit() else line_no,
            "function_name":   _get(row, idx, "function name", "function", "func name", "funcname"),
            "approved_fix":    approved_fix,
            "has_approved_fix": bool(approved_fix),
            "client_accepted": False,
            "source_context":  source_ctx,
        })

    if skipped:
        print(f"  Skipped {skipped} warnings - source files not uploaded")

    warnings.sort(key=lambda w: (-w["severity_rank"], w["file_path"], w["line_start"] or 0))
    return warnings


def parse_report(xlsx_path: Path, source_dir: Path) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(str(xlsx_path), read_only=True, data_only=True)

    # For multi-sheet files, prefer the QAC analysis sheet (Sheet 1)
    ws = wb.active
    for name in wb.sheetnames:
        nl = name.lower()
        if any(k in nl for k in ("analysis", "qac_analysis", "appl", "report")):
            ws = wb[name]
            break

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel sheet is empty.")

    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    fmt = _detect_format(headers)

    print(f"  Format: {fmt.upper()}  |  Sheet: {ws.title!r}")

    uploaded_stems = {
        f.stem.lower()
        for f in list(source_dir.glob("*.c")) + list(source_dir.glob("*.h"))
    }

    return _parse_qac(rows, headers, source_dir, uploaded_stems) if fmt == "qac"         else _parse_mock(rows, headers, source_dir, uploaded_stems)


def main() -> int:
    if len(sys.argv) < 3:
        print("Usage: python parse_polyspace.py <report.xlsx> <source_dir> [out.json]")
        return 2

    xlsx_path  = Path(sys.argv[1])
    source_dir = Path(sys.argv[2])
    out_path   = Path(sys.argv[3]) if len(sys.argv) > 3 else None

    print(f"Parsing {xlsx_path.name} ...")
    warnings = parse_report(xlsx_path, source_dir)
    print(f"  Matched {len(warnings)} warnings")

    if out_path:
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(
            json.dumps({"warning_count": len(warnings), "warnings": warnings},
                       indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
        print(f"  Output: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())