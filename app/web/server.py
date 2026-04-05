"""
app/web/server.py  —  MISRA GenAI Flask Web UI  (Results Viewer)

Architecture:
  - The pipeline (Phase 6a -> 6b -> 7 -> 8) runs ONLY via CLI (orchestrator.py)
  - This web server is a pure results viewer — no LLM, no pipeline, no heavy work
  - It reads evaluated_fixes.json from data/output/<run_id>/ and serves it

Routes:
  GET  /                        -> upload page (index.html)
  POST /api/analyse             -> saves uploads, launches orchestrator.py subprocess
  GET  /api/progress/<job_id>   -> SSE stream of pipeline stdout
  GET  /results/<run_id>        -> results viewer page
  GET  /api/result/<run_id>     -> JSON results from saved evaluated_fixes.json
  GET  /api/runs                -> list all completed run directories
  POST /api/commit              -> returns patched code + download URL
  GET  /api/download/<filename> -> download patched file

Run from project root:
    python app/web/server.py
Then open http://127.0.0.1:5000
"""

import json
import os
import queue
import re as _re
import shutil
import subprocess
import sys
import threading
import time
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

import pandas as pd
from flask import (Flask, Response, jsonify, render_template,
                   request, stream_with_context, send_file)
from flask_cors import CORS
from werkzeug.utils import secure_filename

# ---------------------------------------------------------------------------
# Path setup
# ---------------------------------------------------------------------------
PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

try:
    from app.config.settings import OUTPUT_DIR, CACHE_PATH, DEFAULT_BATCH_SIZE
except ImportError:
    DATA_DIR           = PROJECT_ROOT / "data"
    OUTPUT_DIR         = DATA_DIR / "output"
    CACHE_PATH         = DATA_DIR / "cache" / "results_cache.db"
    DEFAULT_BATCH_SIZE = 5

UPLOAD_DIR = PROJECT_ROOT / "data" / "input" / "web_uploads"
UPLOAD_DIR.mkdir(parents=True, exist_ok=True)
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# Audit Excel — saved to Output_excel_after_run/ inside the project root.
# This resolves to the same absolute folder regardless of where server.py is
# run from, and works on both Windows and Linux without hard-coded user paths.
#
# On your machine this resolves to:
#   C:\Users\sanjay.ravichander\misra_genai_system\misra_genai_system\Output_excel_after_run\audit_report.xlsx
#
AUDIT_EXCEL = PROJECT_ROOT / "Output_excel_after_run" / "audit_report.xlsx"
AUDIT_EXCEL.parent.mkdir(parents=True, exist_ok=True)   # create folder at startup

ALLOWED_EXCEL = {".xlsx", ".xls"}
ALLOWED_C     = {".c", ".h"}

# ---------------------------------------------------------------------------
# Flask app
# ---------------------------------------------------------------------------
_WEB_DIR      = Path(__file__).resolve().parent
_TEMPLATE_DIR = _WEB_DIR / "templates"
_STATIC_DIR   = _WEB_DIR / "static"

if not _TEMPLATE_DIR.exists():
    raise RuntimeError(f"\n\n[server.py] templates/ not found at:\n  {_TEMPLATE_DIR}\n")
if not _STATIC_DIR.exists():
    raise RuntimeError(f"\n\n[server.py] static/ not found at:\n  {_STATIC_DIR}\n")

app = Flask(
    __name__,
    template_folder=str(_TEMPLATE_DIR),
    static_folder=str(_STATIC_DIR),
)
CORS(app)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024

# In-memory job tracker (pipeline subprocess jobs)
JOBS: dict[str, dict] = {}

# ---------------------------------------------------------------------------
# Config — Excel MISRA rule configuration
# ---------------------------------------------------------------------------
_CONFIG_TMP = PROJECT_ROOT / "data" / "_config_tmp"
_CONFIG_TMP.mkdir(parents=True, exist_ok=True)

# token -> saved Excel path
_CONFIG_FILES: Dict[str, Path] = {}

# Folder where user specification Excel files are saved after Apply Configuration
USER_SPEC_FOLDER = PROJECT_ROOT / "data" / "user_specification_excel_folder"
USER_SPEC_FOLDER.mkdir(parents=True, exist_ok=True)

# Locked original — NEVER written to by the app
ORIGINAL_SPEC = PROJECT_ROOT / "data" / "user_specification.xlsx"

# Single working copy saved under user_specification_excel_folder — updated on every Apply
WORKING_SPEC  = USER_SPEC_FOLDER / "user_specification.xlsx"


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    return str(value).strip()


def _normalize_user_category(value: Any) -> str:
    raw = _safe_text(value).upper()
    if not raw:
        return ""
    if raw in {"M", "R", "A"}:
        return raw
    for ch in raw:
        if ch in {"M", "R", "A"}:
            return ch
    return ""


def _display_misra_category(value: Any) -> str:
    raw = _safe_text(value)
    upper = raw.upper()
    if upper == "MISRA-M":
        return "Mandatory"
    if upper == "MISRA-R":
        return "Required"
    if upper == "MISRA-A":
        return "Advisory"
    return raw


def _ensure_user_category_col(df: pd.DataFrame) -> pd.DataFrame:
    df.columns = [str(c).strip() for c in df.columns]
    if "User Category" not in df.columns:
        df["User Category"] = ""
    return df


def _read_excel_df(path: Path) -> pd.DataFrame:
    ext = path.suffix.lower()
    if ext == ".xls":
        try:
            return pd.read_excel(path, engine="xlrd")
        except Exception:
            return pd.read_excel(path)
    return pd.read_excel(path, engine="openpyxl")


def _excel_rows_from_df(df: pd.DataFrame) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    for idx, row in df.iterrows():
        rule_raw = _safe_text(row.get("MISRA Rule", ""))
        rule_list = rule_raw.replace("Rule-", "").replace("Rule_", "").strip()
        rows.append({
            "row_index": int(idx),
            "sl_no": _safe_text(row.get("SI.No", idx + 1)),
            "rule_list": rule_list,
            "misra_category": _safe_text(row.get("MISRA Category", "")),
            "misra_category_display": _display_misra_category(row.get("MISRA Category", "")),
            "user_category": _normalize_user_category(row.get("User Category", "")),
            "warning_message_nos": _safe_text(row.get("Warning Message Nos.", "")),
        })
    return rows


def _get_config_path(token: str) -> Optional[Path]:
    token = Path(token).name
    path = _CONFIG_FILES.get(token)
    if path and path.exists():
        return path
    matches = list(_CONFIG_TMP.glob(f"{token}_*.xlsx"))
    if matches:
        return matches[0]
    matches = list(_CONFIG_TMP.glob(f"{token}_*.xls"))
    if matches:
        return matches[0]
    return None


# ---------------------------------------------------------------------------
# Routes — pages
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Model pre-warm  — loads Mistral-7B into RAM in background at server start
# so the first analysis request doesn't wait 30-60s for model loading.
# ---------------------------------------------------------------------------
def _prewarm_model():
    """Load the LLM into RAM once at startup — runs in a daemon thread."""
    try:
        import sys as _sys
        _sys.path.insert(0, str(PROJECT_ROOT))
        from app.config.settings import LOCAL_MODEL_PATH, LLM_N_CTX, LLM_N_THREADS, LLM_N_GPU_LAYERS
        from app.generation.generate_misra_response import GenerationConfig, LocalLlamaRuntime
        cfg = GenerationConfig(
            model_path=LOCAL_MODEL_PATH,
            n_ctx=LLM_N_CTX,
            n_threads=LLM_N_THREADS,
            n_gpu_layers=LLM_N_GPU_LAYERS,
        )
        print(f"[server] Pre-warming model: {LOCAL_MODEL_PATH}", flush=True)
        LocalLlamaRuntime.get_instance(cfg)
        print("[server] Model ready — first analysis will start instantly", flush=True)
    except Exception as exc:
        print(f"[server] Model pre-warm failed (will load on first run): {exc}", flush=True)

_prewarm_thread = threading.Thread(target=_prewarm_model, daemon=True, name="model-prewarm")
_prewarm_thread.start()

@app.route("/")
def index():
    return render_template("index.html", default_batch=DEFAULT_BATCH_SIZE)


@app.route("/results/<run_id>")
def results(run_id):
    return render_template("results.html", run_id=run_id)


# ---------------------------------------------------------------------------
# Route — list all completed runs
# ---------------------------------------------------------------------------
@app.route("/api/runs")
def list_runs():
    runs = []
    if not OUTPUT_DIR.exists():
        return jsonify(runs=[])
    for d in sorted(OUTPUT_DIR.iterdir(), reverse=True):
        if not d.is_dir():
            continue
        result_file = d / "evaluated_fixes.json"
        if not result_file.exists():
            result_file = d / "fix_suggestions.json"
        if result_file.exists():
            try:
                data         = json.loads(result_file.read_text(encoding="utf-8"))
                results_list = data.get("results", [])
                total  = len(results_list)
                manual = sum(
                    1 for r in results_list
                    if (r.get("evaluation") or r.get("evaluator_result") or {}).get("needs_manual_review")
                )
                runs.append({
                    "run_id": d.name,
                    "total":  total,
                    "manual": manual,
                    "file":   result_file.name,
                    "mtime":  result_file.stat().st_mtime,
                })
            except Exception:
                pass
    return jsonify(runs=runs)


# ---------------------------------------------------------------------------
# Route — get result for a run (reads from saved JSON file)
# ---------------------------------------------------------------------------
@app.route("/api/result/<run_id>")
def get_result(run_id):
    run_id  = secure_filename(run_id)
    run_dir = OUTPUT_DIR / run_id

    if not run_dir.exists():
        job = next((j for j in JOBS.values() if j.get("run_id") == run_id), None)
        if job:
            return jsonify(error="Pipeline still running", status=job["status"]), 202
        return jsonify(error="Run not found"), 404

    # Prefer evaluated_fixes.json, fall back to fix_suggestions.json
    result_file = run_dir / "evaluated_fixes.json"
    if not result_file.exists():
        result_file = run_dir / "fix_suggestions.json"
    if not result_file.exists():
        job = next((j for j in JOBS.values() if j.get("run_id") == run_id), None)
        if job and job["status"] == "running":
            return jsonify(error="Pipeline still running", status="running"), 202
        return jsonify(error="Results not ready yet", status="pending"), 202

    try:
        data = json.loads(result_file.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify(error=f"Failed to read results: {e}"), 500

    results_list = data.get("results", [])

    # Merge source_context + raw warning fields from enriched_warnings.json
    # (fix_suggestions.json does not carry source code — it lives in enriched)
    enriched_path = run_dir / "enriched_warnings.json"
    if enriched_path.exists():
        try:
            enriched_data = json.loads(enriched_path.read_text(encoding="utf-8"))
            enriched_by_id = {
                str(w.get("warning_id")): w
                for w in enriched_data.get("warnings", [])
            }
            for r in results_list:
                wid = str(r.get("warning_id", ""))
                ew  = enriched_by_id.get(wid)
                if ew:
                    # Attach source_context so the UI can show violated code
                    r["source_context"] = ew.get("source_context", {})
                    # Attach raw warning fields (file, line, message, severity)
                    for field in ("file_path", "line_start", "line_end",
                                  "message", "severity", "rule_id",
                                  "function_name", "checker_name",
                                  "category"):   # category = "MISRA-R (Required)" etc. for filter
                        if field not in r and ew.get(field):
                            r[field] = ew[field]
        except Exception:
            pass  # enriched merge is best-effort

    # Build summary
    high = medium = low = manual = cached = 0
    for r in results_list:
        ev   = r.get("evaluation") or r.get("evaluator_result") or {}
        conf = str(ev.get("overall_confidence", r.get("overall_confidence", ""))).lower()
        if conf == "high":     high   += 1
        elif conf == "medium": medium += 1
        else:                  low    += 1
        if ev.get("needs_manual_review") or ev.get("manual_review_required"):
            manual += 1
        if r.get("_from_cache"):
            cached += 1

    summary = {
        "total":  len(results_list),
        "high":   high,
        "medium": medium,
        "low":    low,
        "manual": manual,
        "cached": cached,
    }

    return jsonify({
        "run_id":   run_id,
        "status":   "done",
        "summary":  summary,
        "warnings": results_list,
        "out_dir":  str(run_dir),
    })


# ---------------------------------------------------------------------------
# Route — commit fix (returns patched code as downloadable file)
# ---------------------------------------------------------------------------
@app.route("/api/commit", methods=["POST"])
def commit_fix():
    body       = request.get_json(force=True) or {}
    warning_id = str(body.get("warning_id", "unknown"))
    run_id     = str(body.get("run_id", ""))
    patched    = body.get("patched_code", "")

    if not patched or patched.strip() == "[fix code not available]":
        return jsonify(error="No patched code to commit"), 400

    commit_dir = PROJECT_ROOT / "data" / "commits"
    commit_dir.mkdir(parents=True, exist_ok=True)

    # ── Try to merge patch into the full original source file ──
    full_patched = None
    original_filename = None
    src_file_path = None
    patch_line_start = None   # 1-indexed line number where the fix was applied
    patch_line_count_val = 0  # how many lines were actually changed

    if run_id:
        try:
            run_dir = OUTPUT_DIR / secure_filename(run_id)
            enriched_path = run_dir / "enriched_warnings.json"
            if enriched_path.exists():
                enriched = json.loads(enriched_path.read_text(encoding="utf-8"))
                warnings_list = enriched.get("warnings", [])
                w = next((x for x in warnings_list if str(x.get("warning_id")) == warning_id), None)
                if w:
                    rel_file   = w.get("file_path", "")
                    line_start = int(w.get("line_start") or 0)
                    line_end   = int(w.get("line_end") or line_start)
                    sc         = w.get("source_context", {})
                    ctx_start  = int(sc.get("context_start_line", line_start) if isinstance(sc, dict) else line_start)
                    ctx_end    = int(sc.get("context_end_line",   line_end)   if isinstance(sc, dict) else line_end)
                    fname_only = Path(rel_file).name if rel_file else ""

                    # Search candidate directories for the source file (most specific first)
                    search_roots = []
                    # 1) Exact job dir for THIS run — run_id format is "20240101_120000_<job_id>"
                    #    The job_id is the last segment after the final underscore.
                    exact_job_id = run_id.split("_")[-1] if run_id else ""
                    if exact_job_id and UPLOAD_DIR.exists():
                        exact_job_dir = UPLOAD_DIR / exact_job_id
                        if exact_job_dir.exists():
                            search_roots.append(exact_job_dir / "source_code")
                            search_roots.append(exact_job_dir)
                    # 2) All other web_uploads dirs (newest first — fallback)
                    if UPLOAD_DIR.exists():
                        for d in sorted(UPLOAD_DIR.iterdir(), reverse=True):
                            if d.is_dir() and d.name != exact_job_id:
                                search_roots.append(d / "source_code")
                                search_roots.append(d)
                    # 3) Legacy _upload_tmp dirs
                    _tmp = PROJECT_ROOT / "data" / "_upload_tmp"
                    if _tmp.exists():
                        for d in sorted(_tmp.iterdir(), reverse=True):
                            if d.is_dir():
                                search_roots.append(d / "source")
                                search_roots.append(d)

                    if fname_only:
                        for root in search_roots:
                            candidate = root / fname_only
                            if candidate.exists():
                                src_file_path = candidate
                                original_filename = fname_only
                                break

                    if src_file_path:
                        orig_lines = src_file_path.read_text(encoding="utf-8", errors="replace").splitlines()
                        import re as _re2

                        # ── Helper: preserve original line indentation ──────────────
                        def _apply_indent(orig_ln, new_code):
                            indent = _re2.match(r"^(\s*)", orig_ln).group(1)
                            return indent + new_code.lstrip()

                        # ── Helper: clean prose-style patched_code ──────────────────
                        # LLM sometimes returns "replace X; with Y;" instead of code.
                        # Extract the last C statement (after "with ") as the fix.
                        def _clean_patched(raw):
                            # "replace printf(...); with printf(...);"  → "printf(...);"
                            m_with = _re2.search(r"\bwith\s+(.*)", raw, _re2.IGNORECASE | _re2.DOTALL)
                            if m_with:
                                candidate = m_with.group(1).strip()
                                # If it still has multiple statements, take the last
                                stmts = [s.strip() for s in _re2.split(r";", candidate) if s.strip()]
                                return (stmts[-1] + ";") if stmts else raw
                            # If multiple semicolons, the last statement is the fix
                            stmts = [s.strip() for s in _re2.split(r";", raw) if s.strip()]
                            return (stmts[-1] + ";") if len(stmts) > 1 else raw

                        # ── Helper: structural + token score ───────────────────────
                        # Uses ALL tokens (including C type keywords) for matching,
                        # plus a structural bonus when both patch and source line are
                        # the same syntactic form (declaration vs assignment vs call).
                        def _score_line(src_ln, patch_code):
                            # Token overlap — ALL identifiers, including uint8_t/uint16_t etc.
                            all_patch_toks = set(_re2.findall(r"[a-zA-Z_][a-zA-Z0-9_]*", patch_code))
                            all_src_toks   = set(_re2.findall(r"[a-zA-Z_][a-zA-Z0-9_]*", src_ln))
                            token_score = len(all_patch_toks & all_src_toks)

                            # Structural bonus (+2): both lines are the SAME syntactic category
                            # declaration:  "type var;"   or  "type var = expr;"
                            # assignment:   "var = expr;"
                            # call:         "func(...);"
                            _DECL  = _re2.compile(r"^\s*[a-zA-Z_]\w*(?:\s*\*)?\s+[a-zA-Z_]\w*\s*(?:=|;)")
                            _ASGN  = _re2.compile(r"^\s*[a-zA-Z_]\w*\s*(?:\[.*?\])?\s*=")
                            _CALL  = _re2.compile(r"^\s*[a-zA-Z_]\w*\s*\(")
                            def _kind(s):
                                s = s.strip()
                                if _DECL.match(s): return "decl"
                                if _ASGN.match(s): return "asgn"
                                if _CALL.match(s): return "call"
                                return "other"
                            struct_bonus = 2 if _kind(patch_code) == _kind(src_ln) else 0

                            return token_score + struct_bonus

                        # ── Parse patched_code lines ────────────────────────────────
                        # The LLM may emit:
                        #   A) "28      small = (uint8_t)total + 300;"  ← explicit line number
                        #   B) "uint16_t small;"                        ← bare code, no number
                        #   C) "replace printf(...); with printf(...);" ← English prose
                        patch_map  = {}   # { 1-indexed line number: new code }
                        bare_lines = []   # code strings stripped of any line-number prefix
                        for ln in patched.splitlines():
                            m = _re2.match(r"^\s*(\d+)\s+(.*)", ln)
                            if m:
                                lnum = int(m.group(1))
                                code = m.group(2)
                                patch_map[lnum] = code
                                bare_lines.append(code)
                            else:
                                bare_lines.append(ln)

                        # ── STRATEGY A: SURGICAL (LLM gave explicit line numbers) ───
                        # Replace only those exact numbered lines, everything else stays.
                        if patch_map:
                            merged = list(orig_lines)
                            for lnum, new_code in patch_map.items():
                                idx = lnum - 1
                                if 0 <= idx < len(merged):
                                    merged[idx] = _apply_indent(merged[idx], new_code)
                            patch_line_start    = min(patch_map.keys())
                            patch_line_count_val = len(patch_map)
                            app.logger.info(f"[commit] SURGICAL: replaced lines {sorted(patch_map.keys())}")

                        # ── STRATEGY B: FUZZY (no line numbers — 1-3 line fix) ──────
                        # Clean prose ("replace X with Y"), then score every line in the
                        # context window using token overlap + structural form bonus.
                        # The highest-scoring line is the violated line to replace.
                        elif len(bare_lines) <= 3:
                            raw_patch  = "\n".join(bare_lines)
                            clean_code = _clean_patched(raw_patch)

                            # Score every non-blank line inside the context window
                            ctx_range = orig_lines[max(0, ctx_start - 1): ctx_end]
                            best_score, best_idx = 0, None
                            for rel_i, src_ln in enumerate(ctx_range):
                                if not src_ln.strip():
                                    continue
                                score = _score_line(src_ln, clean_code)
                                if score > best_score:
                                    best_score, best_idx = score, rel_i

                            if best_idx is not None and best_score > 0:
                                abs_idx = (ctx_start - 1) + best_idx
                                merged  = list(orig_lines)
                                merged[abs_idx] = _apply_indent(orig_lines[abs_idx], clean_code)
                                patch_line_start    = abs_idx + 1
                                patch_line_count_val = 1
                                app.logger.info(
                                    f"[commit] FUZZY: matched line {patch_line_start} "
                                    f"score={best_score} patch={repr(clean_code[:60])}"
                                )
                            else:
                                # No confident match — fall back to warning line_start
                                abs_idx = max(0, (line_start or ctx_start) - 1)
                                merged  = list(orig_lines)
                                merged[abs_idx] = _apply_indent(orig_lines[abs_idx], clean_code)
                                patch_line_start    = abs_idx + 1
                                patch_line_count_val = 1
                                app.logger.warning(
                                    f"[commit] FUZZY fallback (no match): "
                                    f"replaced line {patch_line_start}"
                                )

                        # ── STRATEGY C: BLOCK REPLACE (LLM returned the full fixed block) ─
                        # bare_lines > 3 means the LLM gave the whole context rewritten.
                        # Replace the context window with the bare lines verbatim.
                        else:
                            before = orig_lines[: max(0, ctx_start - 1)]
                            after  = orig_lines[ctx_end:]
                            merged = before + bare_lines + after
                            patch_line_start    = ctx_start
                            patch_line_count_val = len(bare_lines)
                            app.logger.info(f"[commit] BLOCK REPLACE mode: lines {ctx_start}-{ctx_end}")

                        full_patched = "\n".join(merged)
        except Exception as merge_err:
            app.logger.warning(f"Patch merge failed (will save snippet only): {merge_err}")

    # Fall back to saving the snippet alone
    content_to_save = full_patched if full_patched else patched
    is_full_file = full_patched is not None

    safe_wid = secure_filename(warning_id)
    base_name = original_filename or f"warning_{safe_wid}"
    stem = Path(base_name).stem
    fname = f"patched_{stem}_{uuid.uuid4().hex[:6]}.c"
    out_path = commit_dir / fname
    out_path.write_text(content_to_save, encoding="utf-8")

    # Save backup of original source file for potential revert
    if src_file_path and src_file_path.exists():
        try:
            backup_path = commit_dir / f"orig_{safe_wid}_{src_file_path.stem}.bak"
            if not backup_path.exists():
                shutil.copy2(str(src_file_path), str(backup_path))
        except Exception:
            pass

    return jsonify({
        "status":            "ok",
        "warning_id":        warning_id,
        "download_url":      f"/api/download/{fname}",
        "filename":          fname,
        "patched_code":      content_to_save,
        "is_full_file":      is_full_file,
        "original_file":     original_filename or "",
        "audit_updated":     False,
        "audit_path":        str(AUDIT_EXCEL),
        "patch_line_start":  patch_line_start,
        "patch_line_count":  patch_line_count_val if is_full_file else len((patched or "").splitlines()),
        "run_id":            run_id,
    })


# ---------------------------------------------------------------------------
# Route — download patched file
# ---------------------------------------------------------------------------
@app.route("/api/download/<filename>")
def download_file(filename):
    filename  = secure_filename(filename)
    file_path = PROJECT_ROOT / "data" / "commits" / filename
    if not file_path.exists():
        return jsonify(error="File not found"), 404
    return send_file(str(file_path), as_attachment=True, download_name=filename)


# ---------------------------------------------------------------------------
# Route — Save audit Excel (manually triggered from Review Report tab)
# ---------------------------------------------------------------------------
@app.route("/api/save_audit", methods=["POST"])
def save_audit():
    body       = request.get_json(force=True) or {}
    warning_id = str(body.get("warning_id", "unknown"))
    run_id     = str(body.get("run_id", ""))
    patched    = body.get("patched_code", "")

    # Build violated_code from enriched warnings
    violated_code = ""
    if run_id:
        try:
            rdir = OUTPUT_DIR / secure_filename(run_id)
            ep   = rdir / "enriched_warnings.json"
            if ep.exists():
                import json as _jx
                ed = _jx.loads(ep.read_text(encoding="utf-8"))
                for wx in ed.get("warnings", []):
                    if str(wx.get("warning_id")) == str(warning_id):
                        sc = wx.get("source_context", {})
                        violated_code = sc.get("context_text", "") if isinstance(sc, dict) else str(sc)
                        break
        except Exception:
            pass

    ok = _update_audit_excel(warning_id, run_id, violated_code, patched)
    if ok:
        return jsonify({"status": "ok", "audit_path": str(AUDIT_EXCEL)})
    return jsonify({"status": "error", "message": "Failed to update audit Excel"}), 500


# ---------------------------------------------------------------------------
# Route — Revert committed fix (restore original source from backup)
# ---------------------------------------------------------------------------
@app.route("/api/revert", methods=["POST"])
def revert_fix():
    body       = request.get_json(force=True) or {}
    warning_id = str(body.get("warning_id", ""))
    run_id     = str(body.get("run_id", ""))

    if not warning_id:
        return jsonify(error="No warning_id provided"), 400

    safe_wid   = secure_filename(warning_id)
    commit_dir = PROJECT_ROOT / "data" / "commits"

    # Find backup files for this warning
    backups = list(commit_dir.glob(f"orig_{safe_wid}_*.bak"))
    if not backups:
        return jsonify(error="No backup found for this warning — cannot revert"), 404

    backup_path = backups[0]
    # Derive original stem from backup filename: orig_<wid>_<stem>.bak
    stem = backup_path.stem[len(f"orig_{safe_wid}_"):]

    # Search upload dirs for the live source file
    exact_job_id = run_id.split("_")[-1] if run_id else ""
    search_roots = []
    if exact_job_id and UPLOAD_DIR.exists():
        d = UPLOAD_DIR / exact_job_id
        if d.exists(): search_roots += [d / "source_code", d]
    if UPLOAD_DIR.exists():
        for d in sorted(UPLOAD_DIR.iterdir(), reverse=True):
            if d.is_dir() and d.name != exact_job_id:
                search_roots += [d / "source_code", d]
    _tmp = PROJECT_ROOT / "data" / "_upload_tmp"
    if _tmp.exists():
        for d in sorted(_tmp.iterdir(), reverse=True):
            if d.is_dir(): search_roots += [d / "source", d]

    restored = False
    for root in search_roots:
        for ext in [".c", ".h"]:
            candidate = root / (stem + ext)
            if candidate.exists():
                shutil.copy2(str(backup_path), str(candidate))
                restored = True
                break
        if restored: break

    if not restored:
        return jsonify(error="Could not find original source file to restore"), 404

    return jsonify({"status": "ok", "message": "File reverted to original"})


# ---------------------------------------------------------------------------
# Audit Excel log
# ---------------------------------------------------------------------------
def _update_audit_excel(warning_id: str, run_id: str,
                         violated_code: str, fixed_code: str) -> bool:
    """Upsert a row for warning_id in the audit Excel report.
    Returns True on success, False on failure (failure is also logged).
    The AUDIT_EXCEL folder is guaranteed to exist (created at server startup).
    """
    import openpyxl
    from openpyxl import Workbook

    COLS = ["Warning Number", "Category", "Rule", "Message",
            "File", "Function", "Violated Code", "Fixed Code",
            "Run ID", "Status", "Timestamp"]

    try:
        # Load or create workbook
        if AUDIT_EXCEL.exists():
            wb = openpyxl.load_workbook(str(AUDIT_EXCEL))
            ws = wb.active
            headers = [c.value for c in ws[1]]
            # Add any missing columns to the right
            for col in COLS:
                if col not in headers:
                    ws.cell(row=1, column=len(headers) + 1, value=col)
                    headers.append(col)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "MISRA Audit"
            for i, h in enumerate(COLS, 1):
                ws.cell(row=1, column=i, value=h)
            headers = list(COLS)

        def col_idx(name):
            try:
                return headers.index(name) + 1
            except ValueError:
                return None

        wid_col = col_idx("Warning Number")
        target_row = None
        if wid_col:
            for row in ws.iter_rows(min_row=2):
                if str(row[wid_col - 1].value) == str(warning_id):
                    target_row = row[0].row
                    break

        # Pull extra fields from enriched_warnings if available
        run_dir       = OUTPUT_DIR / run_id
        enriched_path = run_dir / "enriched_warnings.json"
        ew = {}
        if enriched_path.exists():
            import json as _json
            edata = _json.loads(enriched_path.read_text(encoding="utf-8"))
            for w in edata.get("warnings", []):
                if str(w.get("warning_id")) == str(warning_id):
                    ew = w
                    break

        import datetime
        row_data = {
            "Warning Number": warning_id,
            "Category":       ew.get("severity", ""),
            "Rule":           ew.get("rule_id", ""),
            "Message":        ew.get("message", ""),
            "File":           ew.get("file_path", ""),
            "Function":       ew.get("function_name", ""),
            "Violated Code":  violated_code,
            "Fixed Code":     fixed_code,
            "Run ID":         run_id,
            "Status":         "Fixed",
            "Timestamp":      datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }

        if target_row:
            for col_name, value in row_data.items():
                ci = col_idx(col_name)
                if ci:
                    ws.cell(row=target_row, column=ci, value=value)
        else:
            new_row = [row_data.get(h, "") for h in headers]
            ws.append(new_row)

        wb.save(str(AUDIT_EXCEL))
        app.logger.info(f"Audit Excel updated → {AUDIT_EXCEL}")
        return True

    except Exception as exc:
        app.logger.error(f"Audit Excel update FAILED for warning {warning_id}: {exc}")
        return False


# ---------------------------------------------------------------------------
# Helper — apply rule config filter to uploaded Excel before analysis
# ---------------------------------------------------------------------------
def _filter_excel_by_rules(src_excel: Path, rule_selected: list,
                             rule_overrides: dict) -> Path:
    """Keep only rows whose Rule matches a selected rule_id.
    Returns path to (possibly filtered) Excel file."""
    if not rule_selected:
        return src_excel   # nothing selected → run all

    import openpyxl
    wb = openpyxl.load_workbook(str(src_excel))
    ws = wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]

    # Find the Rule column
    rule_col = next((i for i, h in enumerate(headers)
                     if "rule" in h), None)
    if rule_col is None:
        return src_excel  # can't filter — pass through

    # ── FIX: normalise both sides so "Rule 10.3" matches selected id "10.3" ──
    # Build a set of normalised selected IDs for robust matching
    def _normalise_rule_id(raw: str) -> str:
        """Strip leading 'Rule ' / 'rule ' prefix and whitespace."""
        return _re.sub(r"(?i)^rule\s*", "", str(raw)).strip()

    selected_set = set(_normalise_rule_id(r) for r in rule_selected)

    # Collect rows to keep (skip header)
    keep_rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        raw = str(row[rule_col] or "").strip()
        # Normalise "Rule 10.3" → "10.3"
        normalised = _normalise_rule_id(raw)
        if normalised in selected_set:
            keep_rows.append(row)

    # ── FIX: guard against empty filter result — return original if nothing matched ──
    # This prevents sending a 0-row Excel to the orchestrator, which would cause
    # the pipeline to parse 0 warnings and the UI to show "All 0 records complete".
    if not keep_rows:
        app.logger.warning(
            f"[filter] No rows matched selected rules {sorted(selected_set)} — "
            f"running unfiltered to avoid 0-record pipeline."
        )
        return src_excel

    # Build filtered workbook
    from openpyxl import Workbook
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = ws.title or "Filtered"
    ws2.append([c.value for c in ws[1]])   # header
    for row in keep_rows:
        # Apply override: patch Category column if override exists
        row = list(row)
        cat_col = next((i for i, h in enumerate(headers)
                        if "category" in h), None)
        if cat_col is not None:
            raw_rule = str(row[rule_col] or "")
            norm = _normalise_rule_id(raw_rule)
            ov = rule_overrides.get(norm)
            if ov:
                label = {"M": "MISRA-M (Mandatory)",
                         "R": "MISRA-R (Required)",
                         "A": "MISRA-A (Advisory)"}.get(ov, row[cat_col])
                row[cat_col] = label
        ws2.append(row)

    filtered_path = src_excel.parent / ("filtered_" + src_excel.name)
    wb2.save(str(filtered_path))
    return filtered_path


# ---------------------------------------------------------------------------
# API — Excel MISRA config load / save
# ---------------------------------------------------------------------------
@app.route("/api/config/load", methods=["GET"])
def load_config():
    """Load MISRA rule config from the locked original (or working copy if it exists).
    On first run: copies locked original → WORKING_SPEC.
    On subsequent runs: reads WORKING_SPEC (preserves previous overrides).
    Original is NEVER written to.
    """
    try:
        # Ensure working copy exists
        if not WORKING_SPEC.exists():
            if not ORIGINAL_SPEC.exists():
                return jsonify({"error":
                    f"Original file not found: {ORIGINAL_SPEC}\n"
                    "Place user_specification.xlsx in:\n"
                    f"  {PROJECT_ROOT / 'data'}"}), 404
            shutil.copy2(str(ORIGINAL_SPEC), str(WORKING_SPEC))
            app.logger.info(f"[config/load] Created working copy: {WORKING_SPEC}")
        else:
            app.logger.info(f"[config/load] Reading working copy: {WORKING_SPEC}")

        file_path = WORKING_SPEC
        df = pd.read_excel(str(file_path), engine="openpyxl")
        df.columns = [str(c).strip() for c in df.columns]
        df = df.fillna("")

        RULE_COL     = "MISRA Rule"
        CAT_COL      = "MISRA Category"
        USER_CAT_COL = "User Category"

        missing = [c for c in [RULE_COL, CAT_COL] if c not in df.columns]
        if missing:
            return jsonify({
                "error": f"Expected columns not found: {missing}. Found: {list(df.columns)}"
            }), 400

        rows = []
        for i, row in df.iterrows():
            rule_raw = str(row.get(RULE_COL, "")).strip()
            cat_raw  = str(row.get(CAT_COL,  "")).strip()
            user_cat = str(row.get(USER_CAT_COL, "")).strip() if USER_CAT_COL in df.columns else ""

            if not rule_raw:
                continue

            rule_display = rule_raw
            m = _re.match(r"Rule[-_](.+)", rule_raw, _re.I)
            if m:
                rule_display = m.group(1).strip()
            else:
                m2 = _re.match(r"Dir[-_](.+)", rule_raw, _re.I)
                if m2:
                    rule_display = "Dir " + m2.group(1).strip()

            cat_display = {"MISRA-M": "Mandatory", "MISRA-R": "Required",
                           "MISRA-A": "Advisory"}.get(cat_raw, cat_raw)

            warn_nos = str(row.get("Warning Message Nos.", "")).strip() \
                if "Warning Message Nos." in df.columns else ""

            rows.append({
                "row_index":              int(i),
                "rule_list":              rule_display,
                "misra_category":         cat_display,
                "misra_category_display": cat_display,
                "user_category":          _normalize_user_category(user_cat),
                "warning_message_nos":    warn_nos,
            })

        if not rows:
            return jsonify({"error": "No valid data rows found in Excel"}), 400

        token = uuid.uuid4().hex
        _CONFIG_FILES[token] = WORKING_SPEC

        return jsonify({"token": token, "rows": rows, "count": len(rows)})

    except Exception:
        import traceback
        print("GET /api/config/load ERROR:\n", traceback.format_exc())
        return jsonify({"error": "Internal server error loading config"}), 500


@app.route("/api/config/save", methods=["POST"])
def api_config_save():
    """Save user category selections back to the user_specification Excel."""
    data    = request.get_json(silent=True) or {}
    token   = _safe_text(data.get("token", ""))
    updates = data.get("updates", [])

    if not token:
        return jsonify(error="Missing config token."), 400

    # Always write to the single working copy — original untouched
    if not WORKING_SPEC.exists():
        return jsonify(error="Working copy not found. Please open the modal first."), 404

    try:
        df = _read_excel_df(WORKING_SPEC)
        df.columns = [str(c).strip() for c in df.columns]

        USER_CAT_COL = "User Category"
        if USER_CAT_COL not in df.columns:
            df[USER_CAT_COL] = "-"

        for item in updates:
            try:
                idx      = int(item.get("row_index"))
                user_cat = _normalize_user_category(item.get("user_category", ""))
                if idx in df.index:
                    df.at[idx, USER_CAT_COL] = user_cat if user_cat else "-"
            except Exception:
                continue

        df.to_excel(str(WORKING_SPEC), index=False, engine="openpyxl")
        app.logger.info(f"[config/save] Overrides saved → {WORKING_SPEC}")

        rows = _excel_rows_from_df(df)

    except Exception as exc:
        return jsonify(error=f"Could not save Excel: {exc}"), 400

    return jsonify(status="updated", token=token, rows=rows,
                   saved_file=WORKING_SPEC.name,
                   saved_path=str(WORKING_SPEC))


# ---------------------------------------------------------------------------
# Route — save uploads once (called when user selects files, before any run)
# ---------------------------------------------------------------------------
@app.route("/api/save_uploads", methods=["POST"])
def save_uploads():
    if "warning_report" not in request.files:
        return jsonify(error="No warning report uploaded"), 400
    excel_file = request.files["warning_report"]
    if Path(excel_file.filename).suffix.lower() not in ALLOWED_EXCEL:
        return jsonify(error="Warning report must be .xlsx or .xls"), 400

    c_files = request.files.getlist("source_files")
    if not c_files or all(f.filename == "" for f in c_files):
        return jsonify(error="No C source files uploaded"), 400

    upload_session_id = str(uuid.uuid4())[:8]
    job_dir = UPLOAD_DIR / upload_session_id
    src_dir = job_dir / "source_code"
    src_dir.mkdir(parents=True, exist_ok=True)

    excel_path = job_dir / secure_filename(excel_file.filename)
    excel_file.save(str(excel_path))

    saved_c = []
    for f in c_files:
        if Path(f.filename).suffix.lower() in ALLOWED_C and f.filename:
            dest = src_dir / secure_filename(f.filename)
            f.save(str(dest))
            saved_c.append(dest.name)

    if not saved_c:
        return jsonify(error="No valid .c / .h files received"), 400

    return jsonify(
        upload_session_id=upload_session_id,
        excel_filename=secure_filename(excel_file.filename),
        c_files=saved_c,
    )


# ---------------------------------------------------------------------------
# Route — start analysis (saves uploads, launches orchestrator subprocess)
# ---------------------------------------------------------------------------
@app.route("/api/analyse", methods=["POST"])
def start_analysis():
    try:
        batch_size = max(1, min(15, int(request.form.get("batch_size", DEFAULT_BATCH_SIZE))))
    except (TypeError, ValueError):
        batch_size = DEFAULT_BATCH_SIZE

    upload_session_id = request.form.get("upload_session_id", "").strip()

    if upload_session_id:
        # ── Fast path: files already saved by /api/save_uploads ──
        session_dir = UPLOAD_DIR / secure_filename(upload_session_id)
        src_dir = session_dir / "source_code"
        if not session_dir.exists():
            return jsonify(error="Upload session not found — please re-upload your files"), 400

        # Find the excel file in the session dir
        excel_candidates = [p for p in session_dir.iterdir()
                            if p.suffix.lower() in ALLOWED_EXCEL]
        if not excel_candidates:
            return jsonify(error="Warning report not found in upload session"), 400
        excel_path = excel_candidates[0]

        # Determine which .c files to run (subset or all)
        run_files = request.form.getlist("run_filenames")
        if run_files:
            saved_c = [secure_filename(n) for n in run_files
                       if (src_dir / secure_filename(n)).exists()]
        else:
            saved_c = [p.name for p in src_dir.iterdir()
                       if p.suffix.lower() in ALLOWED_C]

        if not saved_c:
            return jsonify(error="No matching source files found in upload session"), 400

        # Create a new job_id/run_id but reuse the same src_dir
        job_id = str(uuid.uuid4())[:8]
        run_id = time.strftime("%Y%m%d_%H%M%S") + "_" + job_id

    else:
        # ── Standard path: fresh file upload ──
        if "warning_report" not in request.files:
            return jsonify(error="No warning report uploaded"), 400
        excel_file = request.files["warning_report"]
        if Path(excel_file.filename).suffix.lower() not in ALLOWED_EXCEL:
            return jsonify(error="Warning report must be .xlsx or .xls"), 400

        c_files = request.files.getlist("source_files")
        if not c_files or all(f.filename == "" for f in c_files):
            return jsonify(error="No C source files uploaded"), 400

        job_id  = str(uuid.uuid4())[:8]
        run_id  = time.strftime("%Y%m%d_%H%M%S") + "_" + job_id
        job_dir = UPLOAD_DIR / job_id
        src_dir = job_dir / "source_code"
        src_dir.mkdir(parents=True, exist_ok=True)

        excel_path = job_dir / secure_filename(excel_file.filename)
        excel_file.save(str(excel_path))

        saved_c = []
        for f in c_files:
            if Path(f.filename).suffix.lower() in ALLOWED_C and f.filename:
                dest = src_dir / secure_filename(f.filename)
                f.save(str(dest))
                saved_c.append(dest.name)

        if not saved_c:
            return jsonify(error="No valid .c / .h files received"), 400

    # Apply rule config filter — keep only selected rules, apply overrides
    import json as _json
    try:
        rule_selected  = _json.loads(request.form.get("rule_selected", "[]"))
        rule_overrides = _json.loads(request.form.get("rule_overrides", "{}"))
    except Exception:
        rule_selected  = []
        rule_overrides = {}

    # Optional: resume from a previous run so Phase 7 cache hits skip re-generation
    resume_run_id = request.form.get("resume_run_id", "").strip()

    filtered_excel = _filter_excel_by_rules(excel_path, rule_selected, rule_overrides)
    warnings_filtered = len(rule_selected) > 0
    filtered_count = 0
    if warnings_filtered and filtered_excel != excel_path:
        import openpyxl as _opx
        _wb = _opx.load_workbook(str(filtered_excel), read_only=True)
        filtered_count = max(0, _wb.active.max_row - 1)  # exclude header
        _wb.close()

        # ── FIX: abort early if the filter produced 0 rows ──
        # This gives the user a clear error instead of a silent "0 records" run.
        if filtered_count == 0:
            return jsonify(
                error=(
                    "No warnings matched the selected rules. "
                    "Try selecting different rules or clear the rule filter to run all warnings."
                ),
                filtered_count=0,
                warnings_filtered=True,
            ), 400

    q = queue.Queue()
    JOBS[job_id] = {
        "status":     "running",
        "queue":      q,
        "run_id":     run_id,
        "started_at": time.time(),
    }

    t = threading.Thread(
        target=_run_pipeline_subprocess,
        args=(job_id, run_id, str(filtered_excel), str(src_dir), batch_size),
        kwargs={"resume_run_id": resume_run_id},
        daemon=True,
    )
    t.start()

    return jsonify(job_id=job_id, run_id=run_id, c_files=saved_c,
                   batch_size=batch_size, warnings_filtered=warnings_filtered,
                   filtered_count=filtered_count, resumed=bool(resume_run_id))


# ---------------------------------------------------------------------------
# Route — SSE progress stream
# ---------------------------------------------------------------------------
@app.route("/api/progress/<job_id>")
def progress_stream(job_id):
    if job_id not in JOBS:
        return Response('data: {"error": "Job not found"}\n\n',
                        mimetype="text/event-stream")

    def generate():
        q = JOBS[job_id]["queue"]
        while True:
            try:
                msg = q.get(timeout=30)
                yield f"data: {json.dumps(msg)}\n\n"
                if msg.get("type") in ("done", "error"):
                    break
            except queue.Empty:
                yield 'data: {"type": "heartbeat"}\n\n'

    return Response(
        stream_with_context(generate()),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# Background: run orchestrator.py as subprocess, forward stdout as SSE
# ---------------------------------------------------------------------------
def _run_pipeline_subprocess(
    job_id: str,
    run_id: str,
    excel_path: str,
    src_dir: str,
    batch_size: int,
    resume_run_id: str = "",
) -> None:
    job = JOBS[job_id]
    q   = job["queue"]

    def emit(msg: dict) -> None:
        q.put(msg)

    try:
        python_exe   = sys.executable
        orchestrator = str(PROJECT_ROOT / "app" / "pipeline" / "orchestrator.py")

        cmd = [
            python_exe, orchestrator,
            excel_path, src_dir,
            "--run-id", run_id,
        ]
        # If resuming a previous run, tell the orchestrator to skip already-done phases
        if resume_run_id:
            cmd += ["--resume", resume_run_id]

        # Force UTF-8 + unbuffered output from the subprocess.
        # On Windows, Python stdout is block-buffered when writing to a pipe,
        # which means print() output is held in an 8KB buffer and only reaches
        # the server when the buffer fills or the process ends — causing the UI
        # to show nothing for minutes. PYTHONUNBUFFERED=1 forces line-by-line.
        _env = os.environ.copy()
        _env["PYTHONIOENCODING"] = "utf-8"
        _env["PYTHONUTF8"]       = "1"
        _env["PYTHONUNBUFFERED"] = "1"   # ← critical: forces line-buffered stdout on Windows

        # Emit a neutral status — do NOT activate any phase circle yet.
        # The orchestrator's own print() lines will drive the stepper.
        emit({"type": "status", "label": "Pipeline starting…", "detail": ""})

        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,           # line-buffered pipe read on the server side
            cwd=str(PROJECT_ROOT),
            env=_env,
        )

        total_warnings = 0
        _last_wid      = ""   # tracks the last warning_id seen for fix(es) lines

        for line in proc.stdout:
            line = line.rstrip()
            if not line:
                continue

            # Skip evaluation-progress lines — these look like generation lines
            # ("EVAL_PROGRESS [1/1] Evaluating 2883 ...") and must not create cards.
            if line.startswith("EVAL_PROGRESS") or "EVAL_PROGRESS" in line:
                continue

            # Per-record Phase 8 completion — "EVAL_DONE <wid>"
            # Emitted by evaluate_fixes.py after each record is evaluated.
            if line.startswith("EVAL_DONE"):
                parts = line.split()
                eval_wid = parts[1] if len(parts) > 1 else ""
                if eval_wid:
                    emit({"type": "warning_start", "phase": "8",
                          "warning_id": eval_wid,
                          "label": f"Quality check: {eval_wid}"})
                    emit({"type": "warning_done", "phase": "8",
                          "warning_id": eval_wid,
                          "label": f"Check complete: {eval_wid}"})
                continue

            # Always emit as log line
            emit({"type": "log", "detail": line})

            # Phase transitions — match exact orchestrator print() strings
            # e.g. "Phase 6a — Parsing Polyspace report"
            if "Phase 6a" in line:
                emit({"type": "phase_start", "phase": "6a",
                      "label": "Reading file",
                      "detail": "Loading your files", "progress": 5})

            elif "Phase 6b" in line:
                emit({"type": "phase_done", "phase": "6a",
                      "label": "File read complete", "progress": 15})
                emit({"type": "phase_start", "phase": "6b",
                      "label": "Looking up rules",
                      "detail": "Matching rules to warnings", "progress": 16})

            elif "Phase 7" in line:
                emit({"type": "phase_done", "phase": "6b",
                      "label": "Rule lookup complete", "progress": 35})
                emit({"type": "phase_start", "phase": "7",
                      "label": "Fix suggestions",
                      "detail": "Generating fixes with AI", "progress": 31})

            elif "Phase 8" in line:
                emit({"type": "phase_done", "phase": "7",
                      "label": "Fix suggestions complete", "progress": 65})
                emit({"type": "phase_start", "phase": "8",
                      "label": "Quality check",
                      "detail": "Verifying fix quality", "progress": 66})

            # ── FIX: Robust parsed-warning-count extraction ──
            # Original code only matched "Parsed X warnings" with exact word boundary.
            # Orchestrators may print variants like:
            #   "Parsed 16 warnings — High: 4 ..."
            #   "Parsed 16 MISRA warnings"
            #   "16 warnings parsed"
            #   "Found 16 warnings"
            # The regex below handles all these cases.
            _line_lower = line.lower()
            if "warning" in _line_lower and (
                "parsed" in _line_lower
                or "found" in _line_lower
                or "loaded" in _line_lower
                or "read" in _line_lower
            ):
                _m = _re.search(r'(\d+)\s+(?:misra\s+)?warning', line, _re.IGNORECASE)
                if _m:
                    _candidate = int(_m.group(1))
                    if _candidate > 0:
                        total_warnings = _candidate
                emit({"type": "phase_done", "phase": "6a",
                      "label": "Parsing complete",
                      "detail": line.strip(), "progress": 15,
                      "total": total_warnings})

            # Retrieval lines: "  PS002     Rule 10.3     4 rule(s) retrieved"
            if "rule(s) retrieved" in line:
                parts = line.split()
                wid   = parts[0] if parts else ""
                try:
                    n_rules = line.strip().split("rule(s)")[0].strip().split()[-1]
                except Exception:
                    n_rules = "?"
                # warning_start creates the card; warning_done advances it
                emit({"type": "warning_start", "phase": "6b",
                      "label": f"Looking up rules: {wid}",
                      "detail": line.strip(), "progress": 20,
                      "warning_id": wid})
                emit({"type": "warning_done", "phase": "6b",
                      "label": f"Context: {wid} — {n_rules} rule(s)",
                      "detail": line.strip(), "progress": 25,
                      "warning_id": wid})

            # Generation lines: "  [ 1/16] PS002  Rule 10.3 ..."
            if line.strip().startswith("[") and "/" in line and "]" in line:
                try:
                    inner    = line.strip().lstrip("[").split("]", 1)
                    count    = inner[0].strip()
                    rest     = inner[1].strip().split() if len(inner) > 1 else []
                    wid      = rest[0] if rest else ""
                    rule     = rest[1] if len(rest) > 1 else ""
                    cur, tot = count.split("/")
                    cur_i    = int(cur.strip())
                    tot_i    = int(tot.strip())
                    pct      = 31 + int((cur_i / max(1, tot_i)) * 34)
                    _last_wid = wid  # track for fix(es) line matching

                    # ── FIX: also update total_warnings from generation progress ──
                    # If Phase 6a parsing line was missed/different format,
                    # we can still learn the total from "[1/16]" style lines.
                    if total_warnings == 0 and tot_i > 0:
                        total_warnings = tot_i
                        # Re-emit total so the frontend counter updates
                        emit({"type": "total_update", "total": total_warnings})

                    # Emit warning_start so the UI card appears immediately
                    emit({"type": "warning_start", "phase": "7",
                          "label": f"Processing: {wid} ({rule})",
                          "detail": f"{count} of {tot_i} warnings",
                          "progress": pct, "warning_id": wid,
                          "count": cur_i, "total": tot_i, "pct": pct})
                except Exception:
                    pass

            # Fix result line — "  3 fix(es)  254.2s  ✓" — marks warning complete
            if "fix(es)" in line and ("✓" in line or "✗" in line):
                try:
                    emit({"type": "warning_done", "phase": "7",
                          "label": "Fix generated",
                          "detail": line.strip(), "warning_id": _last_wid or ""})
                except Exception:
                    pass

            # Fix result lines: "       3 fix(es)  254.2s"
            if "fix(es)" in line:
                emit({"type": "detail", "detail": line.strip()})

            # Pipeline complete: "Pipeline complete — 955.7s"
            if "Pipeline complete" in line:
                emit({"type": "phase_done", "phase": "8",
                      "label": "Preparing your report",
                      "detail": "Report ready", "progress": 95})

        proc.wait()

        if proc.returncode == 0:
            job["status"] = "done"
            emit({"type": "done", "label": "Analysis complete",
                  "detail": f"Results saved — run ID: {run_id}",
                  "progress": 100, "run_id": run_id,
                  "total": total_warnings})   # ← FIX: include final total in done event
        else:
            job["status"] = "error"
            emit({"type": "error", "label": "Pipeline failed",
                  "detail": f"Process exited with code {proc.returncode}",
                  "progress": 0})

    except Exception as exc:
        import traceback
        job["status"] = "error"
        emit({"type": "error", "label": "Pipeline failed",
              "detail": str(exc), "traceback": traceback.format_exc(),
              "progress": 0})


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    print("=" * 60)
    print("  MISRA GenAI Web UI  —  Results Viewer")
    print("  http://127.0.0.1:5000")
    print("  Pipeline runs via CLI:")
    print("    python app/pipeline/orchestrator.py <excel> <src_dir> --run-id <id>")
    print("=" * 60)
    app.run(host="127.0.0.1", port=5000, debug=False, threaded=True)